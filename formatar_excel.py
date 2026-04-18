"""
formatar_excel.py — Le saida/leads_maringá.csv e gera saida/leads_formatado.xlsx.

Formatacao:
  - Colunas: tipo, nome, telefone, instagram, site
  - Cabecalho em negrito com fundo escuro e texto branco
  - Filtros automaticos em todas as colunas
  - Cada categoria com uma cor de fundo diferente
  - Larguras de coluna ajustadas ao conteudo
  - Links clicaveis nas colunas instagram e site

Uso:
  python formatar_excel.py
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Forca UTF-8 no terminal Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Caminhos ──────────────────────────────────────────────────────────────────

PASTA_SAIDA   = Path(__file__).parent / "saida"
ARQUIVO_CSV   = PASTA_SAIDA / "leads_maringá.csv"
ARQUIVO_XLSX  = PASTA_SAIDA / "leads_formatado.xlsx"

# ── Colunas exportadas (ordem final) ─────────────────────────────────────────

COLUNAS = ["tipo", "nome", "telefone", "instagram", "site"]

CABECALHOS = {
    "tipo":      "Tipo",
    "nome":      "Nome",
    "telefone":  "Telefone",
    "instagram": "Instagram",
    "site":      "Site",
}

# ── Cores por categoria (preenchimento da linha) ──────────────────────────────
# Formato ARGB: FF + hex RGB

CORES_CATEGORIA = {
    "Medico Esteta":       "FFD6EAF8",  # azul claro
    "Dermatologista":      "FFD5F5E3",  # verde claro
    "Clinica de Estetica": "FFFDE8E8",  # rosa claro
    "Clinica Medica":      "FFFFF3CD",  # amarelo claro
    "Cirurgiao Plastico":  "FFFCE5CD",  # laranja claro
}

# Cor padrao para categorias nao mapeadas
COR_PADRAO = "FFF5F5F5"

# ── Estilos fixos ─────────────────────────────────────────────────────────────

FILL_HEADER = PatternFill("solid", fgColor="FF2C3E50")   # azul escuro
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
FONT_LINK   = Font(name="Calibri", color="FF1155CC", underline="single", size=10)
FONT_NORMAL = Font(name="Calibri", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

BORDER_THIN = Border(
    bottom=Side(style="thin", color="FFBDC3C7"),
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def fill_para(categoria: str) -> PatternFill:
    cor = CORES_CATEGORIA.get(categoria, COR_PADRAO)
    return PatternFill("solid", fgColor=cor)


def largura_coluna(col: str, series: pd.Series) -> int:
    """Calcula largura ideal: max entre cabecalho e conteudo, com limites."""
    max_conteudo = series.astype(str).str.len().max()
    largura = max(len(CABECALHOS[col]), max_conteudo if pd.notna(max_conteudo) else 10)
    limites = {
        "tipo":      (16, 24),
        "nome":      (30, 55),
        "telefone":  (14, 20),
        "instagram": (25, 45),
        "site":      (30, 60),
    }
    minimo, maximo = limites.get(col, (12, 50))
    return max(minimo, min(int(largura) + 2, maximo))


# ── Principal ─────────────────────────────────────────────────────────────────

def formatar_excel():
    # 1) Leitura e limpeza do CSV
    if not ARQUIVO_CSV.exists():
        print(f"[ERRO] Arquivo nao encontrado: {ARQUIVO_CSV}")
        sys.exit(1)

    df = pd.read_csv(ARQUIVO_CSV, encoding="utf-8-sig")
    print(f"[OK] CSV carregado: {len(df)} linhas")

    # Garante que so as colunas necessarias existam
    for col in COLUNAS:
        if col not in df.columns:
            df[col] = ""

    df = df[COLUNAS].fillna("")

    # Ordena por categoria (mesma ordem de CORES_CATEGORIA) e depois por nome
    ordem_cats = list(CORES_CATEGORIA.keys())
    df["_ordem"] = df["tipo"].apply(
        lambda t: ordem_cats.index(t) if t in ordem_cats else len(ordem_cats)
    )
    df = df.sort_values(["_ordem", "nome"]).drop(columns="_ordem").reset_index(drop=True)

    # 2) Criacao do workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Maringa"

    # 3) Cabecalho
    for col_idx, col in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=CABECALHOS[col])
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    # Congela a linha de cabecalho
    ws.freeze_panes = "A2"

    # 4) Dados
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        categoria = getattr(row, "tipo", "")
        fill      = fill_para(categoria)

        for col_idx, col in enumerate(COLUNAS, start=1):
            valor = getattr(row, col, "") or ""
            cell  = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = fill
            cell.border = BORDER_THIN

            # Links clicaveis para instagram e site
            if col == "instagram" and valor:
                url = valor if valor.startswith("http") else f"https://{valor}"
                cell.hyperlink = url
                cell.value     = valor
                cell.font      = FONT_LINK
                cell.alignment = ALIGN_LEFT

            elif col == "site" and valor:
                cell.hyperlink = valor
                cell.value     = valor
                cell.font      = FONT_LINK
                cell.alignment = ALIGN_LEFT

            else:
                cell.value     = valor
                cell.font      = FONT_NORMAL
                cell.alignment = ALIGN_LEFT

    # 5) Larguras das colunas
    for col_idx, col in enumerate(COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura_coluna(col, df[col])

    # Altura padrao das linhas de dados
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 16

    ws.row_dimensions[1].height = 20  # cabecalho um pouco mais alto

    # 6) Filtros automaticos (abrange cabecalho + dados)
    ultima_col = get_column_letter(len(COLUNAS))
    ultima_lin = len(df) + 1
    ws.auto_filter.ref = f"A1:{ultima_col}{ultima_lin}"

    # 7) Aba de legenda das cores
    ws_leg = wb.create_sheet("Legenda")
    ws_leg.column_dimensions["A"].width = 26
    ws_leg.column_dimensions["B"].width = 18

    ws_leg.cell(row=1, column=1, value="Categoria").font      = FONT_HEADER
    ws_leg.cell(row=1, column=1).fill      = FILL_HEADER
    ws_leg.cell(row=1, column=1).alignment = ALIGN_CENTER
    ws_leg.cell(row=1, column=2, value="Cor").font      = FONT_HEADER
    ws_leg.cell(row=1, column=2).fill      = FILL_HEADER
    ws_leg.cell(row=1, column=2).alignment = ALIGN_CENTER

    contagens = df["tipo"].value_counts()
    for i, (cat, cor_hex) in enumerate(CORES_CATEGORIA.items(), start=2):
        total = contagens.get(cat, 0)
        c1 = ws_leg.cell(row=i, column=1, value=f"{cat}  ({total} leads)")
        c1.font      = FONT_NORMAL
        c1.fill      = PatternFill("solid", fgColor=cor_hex)
        c1.alignment = ALIGN_LEFT

        c2 = ws_leg.cell(row=i, column=2, value="")
        c2.fill      = PatternFill("solid", fgColor=cor_hex)

    # 8) Salvar
    PASTA_SAIDA.mkdir(parents=True, exist_ok=True)
    wb.save(ARQUIVO_XLSX)

    # 9) Resumo
    print(f"\n{'='*50}")
    print("  Resumo por categoria:")
    print(f"  {'-'*42}")
    for cat in CORES_CATEGORIA:
        n         = len(df[df["tipo"] == cat])
        com_ig    = (df[df["tipo"] == cat]["instagram"] != "").sum()
        com_site  = (df[df["tipo"] == cat]["site"] != "").sum()
        com_tel   = (df[df["tipo"] == cat]["telefone"] != "").sum()
        print(f"  {cat:<22} {n:>3} leads | tel:{com_tel} site:{com_site} ig:{com_ig}")
    print(f"  {'-'*42}")
    print(f"  TOTAL: {len(df)} leads")
    print(f"\n  Salvo em: {ARQUIVO_XLSX}")
    print("=" * 50)


if __name__ == "__main__":
    formatar_excel()
